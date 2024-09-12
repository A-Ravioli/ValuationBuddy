import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel_report(
    ticker,
    stock_data,
    financial_statements,
    economic_data,
    dcf_results,
    comp_data,
    comp_ratios,
):
    wb = Workbook()

    # Overview tab
    ws_overview = wb.active
    ws_overview.title = "Overview"
    _create_overview_tab(ws_overview, ticker, stock_data, economic_data)

    # Financial Statements tabs
    for statement in ["income_statement", "balance_sheet", "cash_flow"]:
        for period in ["annual", "quarterly"]:
            ws = wb.create_sheet(
                f"{statement.replace('_', ' ').title()} ({period.title()})"
            )
            df = financial_statements[statement][period]
            _add_dataframe_to_worksheet(ws, df)

    # DCF tab
    ws_dcf = wb.create_sheet("DCF Analysis")
    _create_dcf_tab(ws_dcf, dcf_results)

    # Comps tab
    ws_comps = wb.create_sheet("Comparable Analysis")
    _add_dataframe_to_worksheet(ws_comps, comp_ratios)

    wb.save(f"{ticker}_financial_analysis.xlsx")


def _create_overview_tab(ws, ticker, stock_data, economic_data):
    info = stock_data["info"]
    history = stock_data["history"]

    ws["A1"] = f"{ticker} Overview"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "52 Week High"
    ws["B3"] = info.get("fiftyTwoWeekHigh", "N/A")
    ws["A4"] = "52 Week Low"
    ws["B4"] = info.get("fiftyTwoWeekLow", "N/A")
    ws["A5"] = "Market Cap"
    ws["B5"] = info.get("marketCap", "N/A")
    ws["A6"] = "Industry"
    ws["B6"] = info.get("industry", "N/A")
    ws["A7"] = "Sector"
    ws["B7"] = info.get("sector", "N/A")

    # Add recent price history
    ws["A9"] = "Recent Price History"
    ws["A9"].font = Font(bold=True)
    _add_dataframe_to_worksheet(ws, history.tail(), start_row=10, start_col=1)


def _create_dcf_tab(ws, dcf_results):
    ws["A1"] = "DCF Analysis Results"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Enterprise Value"
    ws["B3"] = dcf_results["enterprise_value"]

    ws["A5"] = "Projected Free Cash Flows"
    for i, fcf in enumerate(dcf_results["projected_fcf"], start=1):
        ws[f"A{5+i}"] = f"Year {i}"
        ws[f"B{5+i}"] = fcf

    ws["A12"] = "Terminal Value"
    ws["B12"] = dcf_results["terminal_value"]


def _add_dataframe_to_worksheet(ws, df, start_row=1, start_col=1):
    for r, row in enumerate(dataframe_to_rows(df, index=True, header=True), start_row):
        for c, value in enumerate(row, start_col):
            ws.cell(row=r, column=c, value=value)
